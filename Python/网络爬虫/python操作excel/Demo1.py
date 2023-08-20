import openpyxl
import os
# 1.新建表格
path="C:\\Users\\Lenovo\\Desktop\\book1.xlsx"
wb=openpyxl.Workbook(path)#使用openpyxl创建工作表
if os.path.exists(path)==False:
    wb.save(path)
wb.close()#关闭工作表

# count=0
# for i in range(10):
#     count+=1
#     path="C:\\Users\\Lenovo\\Desktop\\book"+str(count)+".xlsx"
#     wb=openpyxl.Workbook(path)
#     wb.save(path)

#2.打开已存在的表格
path='C:\\Users\\Lenovo\\Desktop\\book1.xlsx'
bk=openpyxl.load_workbook(path)#打开已有的表格book1

#3.操作sheet
# bk.create_sheet("学生信息表")
# bk.save(path)
if None!=["学生信息表"]:#判断是否存在,不存在则新建
    print("已经存在")
else:
    st1=bk.create_sheet("学生信息表")
sheet1=bk["学生信息表"]
print(sheet1)
bk.save(path)


