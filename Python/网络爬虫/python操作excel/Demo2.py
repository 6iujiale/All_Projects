import openpyxl
import os
# 1.新建表格
path="C:\\Users\\Lenovo\\Desktop\\book1.xlsx"
wb=openpyxl.Workbook(path)#使用openpyxl创建工作表
if os.path.exists(path)==False:
    wb.save(path)
bk=openpyxl.load_workbook(path)
# bk.create_sheet("学生信息表")
# bk.save(path)
sheet1=bk["学生信息表"]
sheet1.cell(2,1).value="name"
bk.save(path)
'''
# sheet1.cell(2,1).value="name"
sheet1.cell(1,1).value="序号"
# sheet1['C3'].value='Mike'
for i in range(1,47):
    sheet1.cell(i+1,1).value=i
bk.save(path)
'''
# c_range=sheet1["A1:A46"]
# for rows in c_range:
#     for data in rows:
#         print(data.value,end="\n")

