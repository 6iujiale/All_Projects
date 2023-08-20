import openpyxl
import os
addList=[
    ["001","jack",99,89,82],
    ["002","hello",59,10,80],
    ["005","windy",69,89,77],
]
path="C:\\Users\\Lenovo\\Desktop\\ex6.xlsx"
bk=openpyxl.load_workbook(path)#使用openpyxl创建工作表
ss=bk.sheetnames
if "class1" not in ss:
    sheet2=bk.create_sheet("class1")
    bk.save(path)
sheet2=bk["class1"]
tit=["序号","姓名","语文","数学","英语"]
sheet2.append(tit)
for row in addList:
    sheet2.append(row)
bk.save(path)
bk.close()

